from __future__ import annotations

from copy import copy
import os
from pathlib import Path
import tempfile
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries

from lamtools_core.tool import ToolArtifact, ToolCall, ToolResult
from lamtools_core.tool.workspace import relative_workspace_uri, validate_workspace_path


FORMULA_LIMITATION = (
    "Formulas are stored but not calculated by Core; cached results update when the workbook is opened "
    "in a compatible spreadsheet application."
)
FEATURE_LIMITATION = (
    "The structured interface edits cell values, formulas, common formatting, column widths, and freeze panes; "
    "it does not create or edit macros, charts, pivot tables, conditional formatting, or external connections."
)


def _object_schema(properties: dict[str, Any], required: list[str] | None = None) -> dict[str, Any]:
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": properties,
        "required": required or [],
    }


CELL_FORMAT_SCHEMA = _object_schema(
    {
        "bold": {"type": "boolean"},
        "italic": {"type": "boolean"},
        "font_size": {"type": "number", "minimum": 1, "maximum": 409},
        "font_color": {"type": "string", "description": "RGB or ARGB hex color, with optional leading #"},
        "fill_color": {"type": "string", "description": "RGB or ARGB hex color, with optional leading #"},
        "number_format": {"type": "string"},
        "horizontal_alignment": {
            "type": "string",
            "enum": ["general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"],
        },
        "vertical_alignment": {
            "type": "string",
            "enum": ["top", "center", "bottom", "justify", "distributed"],
        },
        "wrap_text": {"type": "boolean"},
    }
)

CELL_SCHEMA = _object_schema(
    {
        "cell": {"type": "string", "description": "Single A1 cell reference"},
        "value": {
            "type": ["string", "number", "boolean", "null"],
            "description": "Literal cell value. Strings beginning with = remain literal text.",
        },
        "formula": {"type": "string", "description": "Excel formula beginning with ="},
        "clear": {"type": "boolean", "description": "Clear the cell value while preserving unspecified formatting"},
        "format": CELL_FORMAT_SCHEMA,
    },
    ["cell"],
)

SHEET_SCHEMA = _object_schema(
    {
        "name": {"type": "string", "minLength": 1, "maxLength": 31},
        "cells": {"type": "array", "items": CELL_SCHEMA, "maxItems": 100_000},
        "column_widths": {
            "type": "array",
            "items": _object_schema(
                {
                    "column": {"type": "string", "description": "Excel column name such as A or BC"},
                    "width": {"type": "number", "minimum": 0, "maximum": 255},
                },
                ["column", "width"],
            ),
            "maxItems": 16_384,
        },
        "freeze_panes": {"type": "string", "description": "Top-left unfrozen cell, such as A2"},
    },
    ["name", "cells"],
)

SPREADSHEET_WRITE_INPUT_SCHEMA = _object_schema(
    {
        "path": {"type": "string", "description": "Destination .xlsx path relative to the workspace"},
        "source_path": {
            "type": "string",
            "description": "Optional existing .xlsx path to edit or copy before applying updates",
        },
        "overwrite": {
            "type": "boolean",
            "description": "Allow replacing an existing destination; editing source_path in place is always allowed",
        },
        "sheets": {"type": "array", "items": SHEET_SCHEMA, "minItems": 1, "maxItems": 100},
    },
    ["path", "sheets"],
)


class SpreadsheetWriteError(ValueError):
    pass


async def write_spreadsheet_tool(call: ToolCall, *, work_root: Path) -> ToolResult:
    args = call.arguments if isinstance(call.arguments, dict) else {}
    try:
        metadata = _write_spreadsheet(args, work_root=work_root)
    except (OSError, SpreadsheetWriteError, ValueError) as exc:
        return ToolResult(
            call_id=call.id,
            name=call.name,
            status="failed",
            error=f"Spreadsheet write failed: {exc}",
        )

    action = str(metadata["action"])
    path = str(metadata["path"])
    content = (
        f"{action.capitalize()} workbook {path}: {metadata['cell_updates']} cell updates across "
        f"{len(metadata['updated_sheets'])} sheet(s).\n"
        f"Limitation: {FORMULA_LIMITATION}\n"
        f"Limitation: {FEATURE_LIMITATION}"
    )
    return ToolResult(
        call_id=call.id,
        name=call.name,
        status="ok",
        content=content,
        artifacts=[
            ToolArtifact(
                kind="file_change",
                uri=path,
                content=None,
                metadata=metadata,
            )
        ],
        metadata=metadata,
    )


def _write_spreadsheet(args: dict[str, Any], *, work_root: Path) -> dict[str, Any]:
    destination = _xlsx_path(args.get("path"), work_root=work_root, label="path")
    source_value = args.get("source_path")
    source = (
        _xlsx_path(source_value, work_root=work_root, label="source_path")
        if source_value not in (None, "")
        else None
    )
    overwrite = args.get("overwrite") is True

    if source is not None:
        if not source.is_file():
            raise SpreadsheetWriteError(f"Source workbook not found: {source_value}")
        if destination.exists() and destination != source and not overwrite:
            raise SpreadsheetWriteError(
                f"Destination already exists: {args.get('path')}; pass overwrite=true to replace it"
            )
        try:
            workbook = load_workbook(source, data_only=False, keep_links=True)
        except Exception as exc:
            raise SpreadsheetWriteError(f"Cannot open source workbook: {exc}") from exc
        action = "edit" if destination == source else "copy"
    else:
        if destination.exists() and not overwrite:
            raise SpreadsheetWriteError(
                f"Destination already exists: {args.get('path')}; pass overwrite=true to replace it"
            )
        workbook = Workbook()
        workbook.remove(workbook.active)
        action = "overwrite" if destination.exists() else "create"

    sheets = args.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise SpreadsheetWriteError("At least one sheet update is required")

    seen_sheets: set[str] = set()
    cell_updates = 0
    formula_updates = 0
    updated_sheets: list[str] = []
    for sheet_payload in sheets:
        if not isinstance(sheet_payload, dict):
            raise SpreadsheetWriteError("Each sheet update must be an object")
        name = _sheet_name(sheet_payload.get("name"))
        if name in seen_sheets:
            raise SpreadsheetWriteError(f"Sheet {name!r} appears more than once")
        seen_sheets.add(name)
        updated_sheets.append(name)
        sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)

        freeze_panes = sheet_payload.get("freeze_panes")
        if freeze_panes not in (None, ""):
            sheet.freeze_panes = _single_cell_reference(freeze_panes, label="freeze_panes")

        for width_payload in sheet_payload.get("column_widths") or []:
            column = _column_name(width_payload.get("column"))
            sheet.column_dimensions[column].width = float(width_payload.get("width"))

        for cell_payload in sheet_payload.get("cells") or []:
            coordinate = _single_cell_reference(cell_payload.get("cell"), label="cell")
            target = sheet[coordinate]
            has_value = "value" in cell_payload
            has_formula = cell_payload.get("formula") is not None
            wants_clear = cell_payload.get("clear") is True
            operations = sum((has_value, has_formula, wants_clear))
            cell_format = cell_payload.get("format") or {}
            if operations > 1:
                raise SpreadsheetWriteError(
                    f"{name}!{coordinate} must use only one of value, formula, or clear"
                )
            if operations == 0 and not cell_format:
                raise SpreadsheetWriteError(f"{name}!{coordinate} has no update")

            if has_value:
                value = cell_payload.get("value")
                target.value = value
                if isinstance(value, str) and value.startswith("="):
                    target.data_type = "s"
            elif has_formula:
                formula = cell_payload.get("formula")
                if not isinstance(formula, str) or not formula.startswith("=") or len(formula) == 1:
                    raise SpreadsheetWriteError(f"Formula for {name}!{coordinate} must begin with =")
                target.value = formula
                formula_updates += 1
            elif wants_clear:
                target.value = None

            if cell_format:
                _apply_cell_format(target, cell_format)
            cell_updates += 1

    if not workbook.sheetnames:
        raise SpreadsheetWriteError("Workbook must contain at least one worksheet")
    if hasattr(workbook, "calculation"):
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

    destination.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{destination.stem}-",
        suffix=".xlsx",
        dir=destination.parent,
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)

    relative_path = relative_workspace_uri(destination, work_root.resolve())
    return {
        "path": relative_path,
        "source_path": relative_workspace_uri(source, work_root.resolve()) if source is not None else "",
        "action": action,
        "size_bytes": destination.stat().st_size,
        "updated_sheets": updated_sheets,
        "sheet_names": list(workbook.sheetnames),
        "cell_updates": cell_updates,
        "formula_updates": formula_updates,
        "limitations": [FORMULA_LIMITATION, FEATURE_LIMITATION],
    }


def _xlsx_path(value: Any, *, work_root: Path, label: str) -> Path:
    if not isinstance(value, str) or not value.strip():
        raise SpreadsheetWriteError(f"Missing {label}")
    try:
        resolved = validate_workspace_path(value, work_root)
    except ValueError as exc:
        raise SpreadsheetWriteError(str(exc)) from exc
    if resolved.suffix.lower() != ".xlsx":
        raise SpreadsheetWriteError(f"{label} must use the .xlsx extension")
    return resolved


def _sheet_name(value: Any) -> str:
    if not isinstance(value, str) or not value.strip():
        raise SpreadsheetWriteError("Sheet name is required")
    name = value.strip()
    if len(name) > 31:
        raise SpreadsheetWriteError(f"Sheet name is longer than 31 characters: {name!r}")
    invalid = set("\\/*?:[]")
    if any(character in invalid for character in name):
        raise SpreadsheetWriteError(f"Sheet name contains an unsupported character: {name!r}")
    return name


def _single_cell_reference(value: Any, *, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise SpreadsheetWriteError(f"{label} must be a single A1 cell reference")
    reference = value.strip().upper()
    try:
        min_column, min_row, max_column, max_row = range_boundaries(reference)
    except ValueError as exc:
        raise SpreadsheetWriteError(f"Invalid {label} reference: {value!r}") from exc
    if min_column != max_column or min_row != max_row or min_column < 1 or min_row < 1:
        raise SpreadsheetWriteError(f"{label} must be a single A1 cell reference")
    return f"{get_column_letter(min_column)}{min_row}"


def _column_name(value: Any) -> str:
    if not isinstance(value, str) or not value.strip():
        raise SpreadsheetWriteError("Column width requires an Excel column name")
    try:
        index = column_index_from_string(value.strip().upper())
    except ValueError as exc:
        raise SpreadsheetWriteError(f"Invalid Excel column name: {value!r}") from exc
    return get_column_letter(index)


def _apply_cell_format(cell: Any, payload: dict[str, Any]) -> None:
    font_keys = {"bold", "italic", "font_size", "font_color"}
    if any(key in payload for key in font_keys):
        font: Font = copy(cell.font)
        if "bold" in payload:
            font.bold = bool(payload["bold"])
        if "italic" in payload:
            font.italic = bool(payload["italic"])
        if "font_size" in payload:
            font.size = float(payload["font_size"])
        if "font_color" in payload:
            font.color = _argb_color(payload["font_color"])
        cell.font = font

    if "fill_color" in payload:
        color = _argb_color(payload["fill_color"])
        cell.fill = PatternFill(fill_type="solid", fgColor=color, bgColor=color)

    alignment_keys = {"horizontal_alignment", "vertical_alignment", "wrap_text"}
    if any(key in payload for key in alignment_keys):
        alignment: Alignment = copy(cell.alignment)
        if "horizontal_alignment" in payload:
            alignment.horizontal = payload["horizontal_alignment"]
        if "vertical_alignment" in payload:
            alignment.vertical = payload["vertical_alignment"]
        if "wrap_text" in payload:
            alignment.wrap_text = bool(payload["wrap_text"])
        cell.alignment = alignment

    if "number_format" in payload:
        cell.number_format = str(payload["number_format"])


def _argb_color(value: Any) -> str:
    if not isinstance(value, str):
        raise SpreadsheetWriteError("Cell colors must be RGB or ARGB hex strings")
    color = value.strip().lstrip("#").upper()
    if len(color) == 6:
        color = "FF" + color
    if len(color) != 8:
        raise SpreadsheetWriteError(f"Invalid color {value!r}; use RGB or ARGB hex")
    try:
        int(color, 16)
    except ValueError as exc:
        raise SpreadsheetWriteError(f"Invalid color {value!r}; use RGB or ARGB hex") from exc
    return color


__all__ = [
    "FEATURE_LIMITATION",
    "FORMULA_LIMITATION",
    "SPREADSHEET_WRITE_INPUT_SCHEMA",
    "SpreadsheetWriteError",
    "write_spreadsheet_tool",
]
